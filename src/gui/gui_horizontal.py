from PyQt6.QtWidgets import (
    QWidget, QLabel, QPushButton, QRadioButton, QFormLayout, QScrollArea, QHBoxLayout,
    QLineEdit, QVBoxLayout, QGridLayout, QStackedWidget, QButtonGroup, QDialog, QTableWidget, QTableWidgetItem, QVBoxLayout
)
from PyQt6.QtCore import Qt
from src.gui import (
    TwoInterval, ThreeInterval, TangentialFourInterval, TangentialFiveInterval, FourInterval, 
    Tangential, Descending, Ascending, Undulant,
    DirectionalProfilesGraphic, HorizontalProfilesGraphic
)
from functools import partial
import pandas as pd
import os
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

direct_profile_dict = {
    0: TwoInterval,
    1: ThreeInterval,
    2: TangentialFourInterval,
    3: TangentialFiveInterval,
    4: FourInterval
}

horiz_profile_dict = {
    0: Tangential,
    1: Descending,
    2: Ascending,
    3: Undulant
}

class Horizontal_Well(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Профиль горизонтальной скважины")
        self.setMinimumSize(900, 700)
        
        # Сразу создаем страницу горизонтального профиля
        layout = QVBoxLayout()
        layout.addWidget(self.horizontal_page())
        self.setLayout(layout) 
    
    def horizontal_page(self):
        # Создаем контейнер и сетку для всего содержимого
        container = QWidget()
        grid = QGridLayout(container)
        
        # --- НАПРАВЛЯЮЩАЯ ЧАСТЬ ---
        
        grid.addWidget(QLabel("Выберите вид направляющей части горизонтальной скважины"), 0, 0, 1, 2)

        # Радиокнопки выбора типа направляющей части
        self.cb_h1 = QRadioButton("1 – двухинтервальный")
        self.cb_h2 = QRadioButton("2 – трёхинтервальный")
        self.cb_h3 = QRadioButton("3 – четырёхинтервальный\n  – танг. участок")
        self.cb_h4 = QRadioButton("4 – пятиинтервальный\n  – танг. участок")
        self.cb_h5 = QRadioButton("5 – четырёхинтервальный")

        radios = [self.cb_h1, self.cb_h2, self.cb_h3, self.cb_h4, self.cb_h5]

        # Добавляем радиокнопки в разметку
        for i, rb in enumerate(radios, start=1):
            grid.addWidget(rb, i, 0)

        # Группируем радиокнопки для взаимно исключающего выбора
        self.radio_group = QButtonGroup()
        for i, rb in enumerate(radios):
            self.radio_group.addButton(rb, i)
        self.cb_h1.setChecked(True)  # по умолчанию выбран первый тип

        # Создаём QStackedWidget с 5 страницами для полей ввода параметров направляющей части
        self.params_stack = QStackedWidget()

        # Для каждого типа создаём страницу с соответствующим набором полей
        def create_params_page(labels):
            w = QWidget()
            f = QFormLayout(w)
            edits = []
            for label in labels:
                le = QLineEdit()
                le.setFixedWidth(100)
                f.addRow(label + ":", le)
                edits.append(le)
            return w, edits
        
        H = "Проектная глубина профиля, м"
        A = "Смещение профиля на проектной глубине, м"
        a = "Угол на проектной глубине, град"
        R1 = "Радиус кривизны 1-го участка, м"
        a1 = "Зенитный угол в конце 1-ого участка профиля, град"
        R3 = "Радиус кривизны 3-го участка, м"
        a3 = "Зенитный угол в конце 3-ого участка профиля, град"
        R4 = "Радиус кривизны 4-го участка, м"
        a2 = "Зенитный угол в конце 2-ого участка профиля, град"
        R2 = "Радиус кривизны 2-го участка, м"

        # Создаем страницы с полями для разных типов направляющей части
        page1, self.inputs_h1 = create_params_page([H, A, a])  
        page2, self.inputs_h2 = create_params_page([H, A, a, a1, R1])                   
        page3, self.inputs_h3 = create_params_page([H, A, a, a1, R1, R3])       
        page4, self.inputs_h4 = create_params_page([H, A, a, a1, R1, R3, a3, R4]) 
        page5, self.inputs_h5 = create_params_page([H, A, a, a1, R1, R2, a2])       

        for p in [page1, page2, page3, page4, page5]:
            self.params_stack.addWidget(p)

        grid.addWidget(self.params_stack, 1, 1, 6, 1)

        # --- ГОРИЗОНТАЛЬНАЯ ЧАСТЬ ---
        
        # Профиль ствола
        grid.addWidget(QLabel("Выберите виды профиля горизонтального ствола"), 7, 0, 1, 2)

        self.rb_straight = QRadioButton("0 – тангенциальный")
        self.rb_desc = QRadioButton("-1 – нисходящий")
        self.rb_asc = QRadioButton("1 – восходящий")
        self.rb_wave = QRadioButton("2 – волнообразный")

        stem_radios = [self.rb_straight, self.rb_desc, self.rb_asc, self.rb_wave]

        for i, rb in enumerate(stem_radios, 8):
            grid.addWidget(rb, i, 0)

        # Группируем радиокнопки горизонтальной части
        self.stem_group = QButtonGroup()
        for i, rb in enumerate(stem_radios):
            self.stem_group.addButton(rb, i)
        self.rb_straight.setChecked(True)  # по умолчанию тангенциальный

        # stacked_stem - страницы с параметрами профиля ствола
        self.stacked_stem = QStackedWidget()
        
        # Создаем страницы для каждого типа горизонтального профиля
        def create_horizontal_params_page(labels):
            w = QWidget()
            f = QFormLayout(w)
            edits = []
            for label in labels:
                le = QLineEdit()
                le.setFixedWidth(120)
                f.addRow(label + ":", le)
                edits.append(le)
            return w, edits

        # Параметры для разных типов горизонтальных профилей
        S_l = "Протяженность горизонтального участка по пласту, м"
        T_down = "Предельное отклонение оси вниз, м"
        T_up = "Предельное отклонение оси вверх, м"
        R_horiz = "Радиус кривизны горизонтального участка, м"

        # Создаем страницы для каждого типа
        h_page1, self.inputs_straight = create_horizontal_params_page([S_l])  # Только протяженность
        h_page2, self.inputs_desc = create_horizontal_params_page([S_l, T_down])  # Протяженность + отклонение вниз
        h_page3, self.inputs_asc = create_horizontal_params_page([S_l, T_up])  # Протяженность + отклонение вверх
        h_page4, self.inputs_wave = create_horizontal_params_page([S_l, T_up, T_down, R_horiz])  # Все параметры

        for hp in [h_page1, h_page2, h_page3, h_page4]:
            self.stacked_stem.addWidget(hp)

        grid.addWidget(self.stacked_stem, 8, 1, 5, 1)

        # Кнопки Назад и Готово
        back = QPushButton("Назад")
        back.clicked.connect(self.close)
        grid.addWidget(back, 13, 0)

        finish = QPushButton("Готово")
        finish.clicked.connect(self.on_confirm_horizontal)
        grid.addWidget(finish, 13, 1, alignment=Qt.AlignmentFlag.AlignRight)

        # --- Оборачиваем все это в QScrollArea ---
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)

        # Связываем переключение радиокнопок со сменой страниц параметров
        self.radio_group.buttonClicked.connect(self.on_horizontal_type_changed)
        self.stem_group.buttonClicked.connect(self.on_stem_type_changed)
        
        # Установим изначально текущие страницы
        self.params_stack.setCurrentIndex(0)
        self.stacked_stem.setCurrentIndex(0)

        return scroll

# Методы для обработки переключения радиокнопок
    def on_horizontal_type_changed(self, button):
        """Обработчик смены типа направляющей части"""
        index = self.radio_group.id(button)
        self.params_stack.setCurrentIndex(index)

    def on_stem_type_changed(self, button):
        """Обработчик смены типа горизонтального профиля"""
        index = self.stem_group.id(button)
        self.stacked_stem.setCurrentIndex(index)

    def on_confirm_horizontal(self):
        
        try:
            idx_direct, vals_direct, idx_horz, vals_horz = self.read_params()
            directional_profile = self.create_profile(idx_direct, vals_direct, direct_profile_dict)
            horizontal_profile = self.create_profile(idx_horz, vals_direct[:3]+vals_horz, horiz_profile_dict)
            table_data = self.build_table_data(directional_profile, horizontal_profile)

            dlg = ResultDialog(table_data, directional_profile, horizontal_profile, self)
            dlg.exec()

        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Ошибка", f"Ошибка при вычислениях:\n{e}")
        
    
    def read_params(self):
        def safe_float(text):
            try:
                return float(text)
            except Exception:
                raise ValueError(f"Некорректное значение: '{text}'")

        idx_nav = self.params_stack.currentIndex()
        if idx_nav == 0:
            vals_nav = [safe_float(le.text()) for le in self.inputs_h1]
        elif idx_nav == 1:
            vals_nav = [safe_float(le.text()) for le in self.inputs_h2]
        elif idx_nav == 2:
            vals_nav = [safe_float(le.text()) for le in self.inputs_h3]
        elif idx_nav == 3:
            vals_nav = [safe_float(le.text()) for le in self.inputs_h4]
        elif idx_nav == 4:
            vals_nav = [safe_float(le.text()) for le in self.inputs_h5]
        else:
            vals_nav = []

        idx_horz = self.stacked_stem.currentIndex()
        if idx_horz == 0:
            vals_horz = [safe_float(le.text()) for le in self.inputs_straight]
        elif idx_horz == 1:
            vals_horz = [safe_float(le.text()) for le in self.inputs_desc]
        elif idx_horz == 2:
            vals_horz = [safe_float(le.text()) for le in self.inputs_asc]
        elif idx_horz == 3:
            vals_horz = [safe_float(le.text()) for le in self.inputs_wave]
        else:
            vals_horz = []

        return idx_nav, vals_nav, idx_horz, vals_horz

    def create_profile(self, idx, vals, profile_dict):
        profile_cls = profile_dict.get(idx)
        if profile_cls:
            return profile_cls(*vals)
        else:
            raise ValueError("Неверный тип профиля")


    def build_table_data(self, directional_profile, horizontal_profile):
        def format_float(val):
            try:
                return f"{val:.2f}" if val is not None else ""
            except Exception:
                return ""

        dp = directional_profile
        hp = horizontal_profile

        rows = []

        # Данные направляющей части
        depths = getattr(dp, 'depth', getattr(dp, 'depths', []))
        lengths = getattr(dp, 'lengths_of_the_bores', [])
        intervals = getattr(dp, 'lengths_of_the_intervals', [])
        offsets = getattr(dp, 'dislocations', [])
        angles = getattr(dp, 'angles', [])
        intensities = getattr(dp, 'intensities', [])

        for i in range(len(depths)):
            rows.append([
                str(i + 1),
                format_float(depths[i] if i < len(depths) else None),
                format_float(lengths[i] if i < len(lengths) else None),
                format_float(intervals[i] if i < len(intervals) else None),
                format_float(offsets[i] if i < len(offsets) else None),
                format_float(angles[i] if i < len(angles) else None),
                format_float(intensities[i] if i < len(intensities) else None),
            ])

        # --- Горизонтальный участок ---

        last_length = lengths[-1] if lengths else 0
        L_h = hp.L_h if hp.L_h is not None else 0

        horiz_total_length = last_length + L_h

        stem_idx = self.stacked_stem.currentIndex()
        if stem_idx == 0:
            zenith_angle = angles[-1] if angles else 0
        else:
            zenith_angle = hp.a_h if hp.a_h is not None else 0

        # Для других параметров берем соответствующие свойтва, с обработкой отсутствия данных
        horiz_depth = hp.H_h if hasattr(hp, 'H_h') else None
        horiz_offset = hp.A_h if hasattr(hp, 'A_h') else None
        horiz_intervals = hp.lengths_of_the_intervals if hasattr(hp, 'lengths_of_the_intervals') else []
        horiz_intensities = hp.intensities if hasattr(hp, 'intensities') else []

        interval_len = horiz_intervals[0] if len(horiz_intervals) > 0 else None
        intensity = horiz_intensities[0] if len(horiz_intensities) > 0 else None

        rows.append([
            "Горизонтальный участок",
            format_float(horiz_depth),
            format_float(horiz_total_length),
            format_float(interval_len),
            format_float(horiz_offset),
            format_float(zenith_angle),
            format_float(intensity)
        ])

        return rows

class ResultDialog(QDialog):
    def __init__(self, table_data, directional_profile, horizontal_profile, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Результаты расчёта профиля")
        self.resize(950, 220)
        layout = QVBoxLayout(self)

        row = QHBoxLayout()

        self.table = QTableWidget(len(table_data), 7)
        self.headers = [
            "Номер участка",
            "Глубина по вертикали, м",
            "Длина ствола, м",
            "Длина интервала, м",
            "Смещение, м",
            "Зенитный угол, град.",
            "Интенсивность искривления, град./10м"
        ]
        self.table.setHorizontalHeaderLabels(self.headers)

        for row_idx, row_data in enumerate(table_data):
            for col_idx, val in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

        self.table.resizeColumnsToContents()
        layout.addWidget(self.table)

        btn_excel = QPushButton("Записать результаты в файл")
        btn_excel.clicked.connect(self.on_open_excel)

        btn_graphics = QPushButton("Показать график")
        btn_graphics.clicked.connect(partial(self.on_open_graphics, directional_profile, horizontal_profile))
        row.addWidget(btn_graphics)
        row.addWidget(btn_excel)

        layout.addLayout(row)
        self.table.horizontalHeader().setDefaultSectionSize(120)     
        self.table.horizontalHeader().setStretchLastSection(True) 
        self._excel_path = "results.xlsx"
        self._table_data = table_data

    def on_open_excel(self):
        # Сначала записываем обычный DataFrame
        df = pd.DataFrame(self._table_data, columns=self.headers)
        try:
            ncols = len(self.headers)
            if os.path.exists(self._excel_path):
                start_row = pd.read_excel(self._excel_path).shape[0] + 2
                with pd.ExcelWriter(self._excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, index=False, header=False, startrow=start_row)
                wb = openpyxl.load_workbook(self._excel_path)
                ws = wb.active
            else:
                with pd.ExcelWriter(self._excel_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False)
                wb = openpyxl.load_workbook(self._excel_path)
                ws = wb.active
                # --- СТИЛИ ДЛЯ ЗАГОЛОВКА ---
                header_fill = PatternFill(fill_type="solid", fgColor="4A90E2")  # синий фон
                header_font = Font(color="FFFFFF", bold=True)
                thick_border = Border(bottom=Side(style="thick", color="000000"))
                center = Alignment(horizontal="center", vertical="center")
                # Стиль для всей строки заголовка
                for col_num in range(1, ncols + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thick_border
                    cell.alignment = center
                    # Ширина столбца (на всю таблицу, можно индивидуально)
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = 40
            row = ws.max_row
            for col_num in range(1, ncols + 1):
                ws.cell(row=2, column=col_num).border = Border(top=Side(style="thick", color="000000"))
                ws.cell(row=row, column=col_num).border = Border(top=Side(style="thick", color="FF9900"), bottom=Side(style="thick", color="000000"))
            wb.save(self._excel_path)
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Ошибка", f"Не удалось открыть файл:\n{e}")

    


    def on_open_graphics(self, dir_prof, horiz_prof):
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        # создаём фигуру и оси
        fig, axes = plt.subplots(figsize=(9, 6))

        # рисуем профили
        directional_profile = DirectionalProfilesGraphic(dir_prof, axes)
        horizontal_profile = HorizontalProfilesGraphic(horiz_prof, axes)
        directional_profile.draw()
        horizontal_profile.draw()

        # создаём канвас для PyQt6
        plt.show()

        




        





